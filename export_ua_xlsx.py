#!/usr/bin/env python3
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

SRC = Path('/home/cri/.openclaw/media/inbound/file_27---00e05754-bc3e-4436-a961-efcac1185b22.xlsx')

@dataclass
class DocScore:
    sheet_name: str
    title: str
    scores: list[int]          # 23 items
    comments: dict[int,str]    # item -> UA comment


def ua_comment(default: str, evidence: str|None=None) -> str:
    if evidence:
        return f"{default}\nДоказ/де в документі: {evidence}"
    return default


def main() -> None:
    wb = openpyxl.load_workbook(SRC)
    tmpl = wb[wb.sheetnames[0]]

    # Prepare single-appraiser scores (Appraiser A) from current report.
    docs: list[DocScore] = [
        DocScore(
            sheet_name='Документ 1 (Onkopedia)',
            title='PNH — Onkopedia (вересень 2024) [HTML+PDF]',
            scores=[6,5,6, 5,3,5, 2,1,2,2,4,3,1,1, 6,6,6, 3,2,2,1, 3,5],
            comments={
                1: ua_comment('Цілі/опис теми наведені на початку.', 'PDF, стор. 5: «Date of document: September 2024» + «Summary»'),
                5: ua_comment('Є ознаки залучення пацієнтських організацій через афіліації (частково).', 'PDF, стор. 34–35: «Aplastische Anämie & PNH e.V.», «Stiftung lichterzellen»'),
                7: ua_comment('Систематичний пошук доказів не описаний явно (стратегія пошуку/БД/дати відсутні).'),
                23: ua_comment('Конфлікти інтересів згадані, але механізм управління/деталі в витягу не розкриті.', 'PDF, стор. 35: «Disclosure of Potential Conflicts of Interest…»'),
            }
        ),
        DocScore(
            sheet_name='Документ 2 (Консенсус)',
            title='Консенсус щодо діагностики та лікування ПНГ (HTCT, 2021) [PDF]',
            scores=[5,4,5, 5,2,4, 2,1,2,3,5,3,1,1, 5,5,5, 4,3,4,2, 2,6],
            comments={
                1: ua_comment('Мета/сфера: консенсус щодо діагностики та лікування ПНГ.', 'PDF, стор. 1–2: назва + опис мети в вступі'),
                20: ua_comment('Ресурсні наслідки розглянуті (вартість/бюджет).', 'PDF, стор. 7: «high cost… funding… public health system (SUS)… public budget…»'),
                23: ua_comment('COI задекларовано.', 'PDF, стор. 7: «The authors declare no conflicts of interest.»'),
            }
        ),
        DocScore(
            sheet_name='Документ 3 (Огляд+Delphi)',
            title='Систематичний огляд + експертна думка (Adv Ther, 2023) [PDF]',
            scores=[4,4,4, 5,1,4, 6,5,5,5,4,5,2,2, 4,4,4, 2,1,1,1, 2,3],
            comments={
                7: ua_comment('Систематичні методи пошуку описані (PRISMA, Embase + PubMed/Medline).', 'PDF, стор. 10'),
                8: ua_comment('Критерії відбору/фільтри та пошукові терміни описані.', 'PDF, стор. 10: «Only full-text… last six years… search terms…»'),
                10: ua_comment('Метод формування рекомендацій описаний (модифікований Delphi).', 'PDF, стор. 10'),
                5: ua_comment('Пацієнти не залучались.', 'PDF, стор. 10: «As no patient was involved…»'),
            }
        ),
    ]

    # Remove any existing extra sheets; keep template only.
    for name in list(wb.sheetnames)[1:]:
        del wb[name]

    # Helper: locate row numbers where item numbers appear in column B.
    def row_for_item(ws, item_no: int) -> int|None:
        for r in range(1, ws.max_row+1):
            v = ws.cell(r, 2).value
            if v is None:
                continue
            try:
                if int(v) == item_no:
                    return r
            except Exception:
                continue
        return None

    # Create per-doc sheets.
    for i, d in enumerate(docs):
        ws = tmpl if i == 0 else wb.copy_worksheet(tmpl)
        ws.title = d.sheet_name

        # Fill title cell (row 2, col A)
        ws.cell(2, 1).value = d.title

        # Fill items
        for item_no, score in enumerate(d.scores, start=1):
            r = row_for_item(ws, item_no)
            if not r:
                continue
            ws.cell(r, 3).value = score
            ws.cell(r, 4).value = d.comments.get(item_no, 'Не описано/не знайдено у витягу з тексту.').strip()

    out = Path('agree2_pnh_single_appraiser_ua.xlsx')
    wb.save(out)
    print('Wrote', out)


if __name__ == '__main__':
    main()
